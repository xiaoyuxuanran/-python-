import openpyxl
MtE_list=['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']
wb = openpyxl.load_workbook('./other/information.xlsx')
sh = wb["学生信息"]
ws=wb.active

def get_min_and_max():
    data_max='a1:'+str(MtE_list[int(ws.max_column)-1])+str(ws.max_row)
    # print("max_column获取的最大列数：",ws.max_column)
    # print("max_row获取的最大行数：",ws.max_row)
    return data_max
list=[[c.value for c in row] for row in sh[str(get_min_and_max())]]
def print_student_data():
    a=str()
    for i in range(1,int(ws.max_row)):
        # print("学号:%d 姓名:%s 年龄:%d 身高:%d"%(list[i][0],list[i][1],list[i][2],list[i][3]))
        a=a+str("学号:%d 姓名:%s 年龄:%d 身高:%d"%(list[i][0],list[i][1],list[i][2],list[i][3]))+'\n'
    return a
def save_excle():
    wb.save('./other/information.xlsx')

def shuaxin():
    wb = openpyxl.load_workbook('./other/information.xlsx')
    sh = wb["学生信息"]
    # wb = openpyxl.load_workbook('./other/log_in.xlsx')
    # sh = wb["账号密码"]
    ws=wb.active
    list=[[c.value for c in row] for row in sh[str(get_min_and_max())]]














# print(print_student_data())

# for i in range(1,10):
#     print("%s:%s %s:%s %s:%s %s:%s"%(list[0][0],list[i][0],list[0][1],list[i][1],list[0][2],list[i][2],list[0][3],list[i][3],))

# def get_min_and_max():
#     ws=wb.active
#     data_max='a1:'+str(MtE_list[int(ws.max_column)-1])+str(ws.max_row)
#     # print("max_column获取的最大列数：",ws.max_column)
#     # print("max_row获取的最大行数：",ws.max_row)
#     return data_max
# get_min_and_max()

# def get_real_max_column(ws):
#     real_max_column = ws.max_column
#     columns = [column for column in ws.columns]
#     while real_max_column>0:
#         column_dict = {c.value for c in columns[real_max_column-1]}
#         if column_dict=={None}:
#             real_max_column = real_max_column-1
#         else:
#             break

#     return real_max_column


# real_max_column = get_real_max_column(ws)
# print("通过自定义函数获得的最大列数：",real_max_column)
