import tkinter
import musicplayer as mp
import time
import openpyxl
import sometest
import easygui
MtE_list=['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']
wb = openpyxl.load_workbook('./other/log_in.xlsx')
sh = wb["账号密码"]
ws=wb.active
def get_min_and_max():
    data_max='a1:'+str(MtE_list[int(ws.max_column)-1])+str(ws.max_row)
    return data_max
list=[[c.value for c in row] for row in sh[str(get_min_and_max())]]
flag1=[0]

class tkinter_ui:
    def __init__(self):
        self.root = tkinter.Tk()
        self.root.wm_attributes('-topmost',1)
        self.x, self.y = 0, 0
        self.window_size = '1000x600'
        self.root.overrideredirect(1)
        self.root.attributes("-alpha", 0.8)
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        ww = 1000
        wh =600
        x = (sw-ww) / 2
        y = (sh-wh) / 2
        self.root.geometry("%dx%d+%d+%d" %(ww,wh,x,y))
        self.root.configure(bg="#F8FF84")
        def compare(x,y):
            x=str(x)
            y=str(y)
            for i in range(ws.max_row):
                if x==str(list[i][0]): 
                    if y==str(list[i][1]):
                        text6.config(text="登录成功")
                        self.root.destroy()
                        time.sleep(1)
                        window = information()
                        window.run()
                        break
                    else:
                        text6.config(text="登录失败")
                else:
                        text6.config(text="登录失败")
        def b1():
            x=E1.get()
            y=E2.get()
            compare(x,y)
        text1 = tkinter.Label(self.root, text="学生信息管理系统",bd="30px",bg='#F8FF84',font=('华文黑体',20,'bold'))
        text1.pack()
        text2= tkinter.Label(self.root,text="用户名：",bd="30px",bg='#F8FF84',font=('华文宋体',20,'bold'))
        text2.place(x=200,y=150,height=50)
        text3= tkinter.Label(self.root,text="密    码：",bd="30px",bg='#F8FF84',font=('华文宋体',20,'bold'))
        text3.place(x=200,y=250,height=50)
        E1=tkinter.Entry(self.root,bg='white',font=('华文宋体',20))
        E1.place(x=400,y=150,height=50,width=300)
        E2=tkinter.Entry(self.root,bg='white',font=('华文宋体',20))
        E2.place(x=400,y=250,height=50,width=300)
        text5= tkinter.Label(self.root,text="登录状态:",bd="30px",bg='#F8FF84',font=('华文宋体',20,'bold'))
        text5.place(x=200,y=350,height=50)
        text6= tkinter.Label(self.root,text="未登录",bd="30px",bg='#F8FF84',font=('华文宋体',20,'bold'))
        text6.place(x=400,y=350,height=50)
        b1=tkinter.Button(self.root,text="登       录",command=b1,bg='#7FF7FF',font=('华文宋体',20,'bold'))
        b1.place(x=250,y=450,height=50,width=500)
        text5= tkinter.Label(self.root,text="注意事项:关闭窗口请双击",bd="30px",bg='#F8FF84',font=('华文黑体',10,'bold'))
        text5.place(x=350,y=580,height=20)
        self.root.bind("<B1-Motion>", self.move)
        self.root.bind("<Button-1>", self.get_point)
        self.root.bind("<Double-Button-1>", self.close)
    def move(self, event):
        new_x = (event.x - self.x) + self.root.winfo_x()
        new_y = (event.y - self.y) + self.root.winfo_y()
        s = f"{self.window_size}+{new_x}+{new_y}"
        self.root.geometry(s)
    def get_point(self, event):
        self.x, self.y = event.x, event.y
    def run(self):
        self.root.mainloop()
    def close(self, event):
        flag1[0]=1
        self.root.destroy()
class information:
    def __init__(self):
        self.root = tkinter.Tk()
        self.x, self.y = 0, 0
        self.window_size = '1000x600'
        self.root.overrideredirect(1)
        self.root.attributes("-alpha", 0.8)
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        ww = 1000
        wh =600
        x = (sw-ww) / 2
        y = (sh-wh) / 2
        self.root.geometry("%dx%d+%d+%d" %(ww,wh,x,y))
        self.root.configure(bg="#F8FF84")
        def b1():
            sometest.shuaxin()
            text4.config(text=sometest.print_student_data())
        def b2():
            box0=easygui.enterbox("请输入要添加的学号")
            box1=easygui.enterbox("请输入要添加的姓名")
            box2=easygui.enterbox("请输入要添加的年龄")
            box3=easygui.enterbox("请输入要添加的身高")
            MAX_line=int(sometest.ws.max_row)+1
            new_data1='A'+str(MAX_line)
            new_data2='B'+str(MAX_line)
            new_data3='C'+str(MAX_line)
            new_data4='D'+str(MAX_line)
            sometest.ws[new_data1]=int(box0)
            sometest.ws[new_data2]=str(box1)
            sometest.ws[new_data3]=int(box2)
            sometest.ws[new_data4]=int(box3)
            sometest.save_excle()
        text1 = tkinter.Label(self.root, text="学生信息管理系统正式界面",bd="30px",bg='#F8FF84',font=('华文黑体',30,'bold'))
        text1.pack()
        b1=tkinter.Button(self.root,text="查  询",command=b1,bg='pink',font=('华文宋体',20,'bold'))
        b1.place(x=100,y=150,height=50,width=100)
        b2=tkinter.Button(self.root,text="增  加",command=b2,bg='pink',font=('华文宋体',20,'bold'))
        b2.place(x=100,y=250,height=50,width=100)
        text4 = tkinter.Label(self.root, text="",bd="30px",bg='black',anchor="w",font=('华文宋体',10,'bold'),fg = 'white')
        text4.place(x=300,y=150,height=250,width=500)
        self.root.bind("<B1-Motion>", self.move)
        self.root.bind("<Button-1>", self.get_point)
        self.root.bind("<Double-Button-1>", self.close)
    def move(self, event):
        new_x = (event.x - self.x) + self.root.winfo_x()
        new_y = (event.y - self.y) + self.root.winfo_y()
        s = f"{self.window_size}+{new_x}+{new_y}"
        self.root.geometry(s)
    def get_point(self, event):
        self.x, self.y = event.x, event.y
    def run(self):
        self.root.mainloop()
    def close(self, event):
        flag1[0]=1
        self.root.destroy()
def run():
    window = tkinter_ui()
    window.run()
